"""
GRC Evidence Platform - Excel Report Generator

This module generates Excel (.xlsx) reports with multiple sheets for:
1. Findings - All evidence findings with status
2. Remediation - Auto-remediation actions and status
3. Compliance - Framework compliance status
4. Summary - Executive summary with key metrics

Author: GRC Platform Team
Version: 2.1
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional
from dataclasses import dataclass


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not available. Install with: pip install openpyxl>=3.1.0")


@dataclass
class ExcelReportConfig:
    """Configuration for Excel report generation."""
    
    # Sheet names
    sheet_findings: str = "Findings"
    sheet_remediation: str = "Remediation"
    sheet_compliance: str = "Compliance"
    sheet_summary: str = "Summary"
    
    # Formatting
    header_font: Dict[str, Any] = None
    header_fill: Dict[str, Any] = None
    header_alignment: Dict[str, Any] = None
    
    def __post_init__(self):
        """Initialize default formatting."""
        if self.header_font is None:
            self.header_font = {
                'name': 'Calibri',
                'size': 11,
                'bold': True,
                'color': 'FFFFFF'
            }
        if self.header_fill is None:
            self.header_fill = {
                'fill_type': 'solid',
                'start_color': '4472C4',
                'end_color': '4472C4'
            }
        if self.header_alignment is None:
            self.header_alignment = {
                'horizontal': 'center',
                'vertical': 'center',
                'wrap_text': True
            }


class ExcelReportGenerator:
    """Generates Excel reports for GRC evidence findings."""
    
    def __init__(self, config: Optional[ExcelReportConfig] = None):
        """
        Initialize Excel report generator.
        
        Args:
            config: Optional configuration for Excel report generation
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required but not installed. "
                "Install with: pip install openpyxl>=3.1.0"
            )
        
        self.config = config or ExcelReportConfig()
        self.workbook = None
        self.worksheet = None
        
    def create_workbook(self) -> 'openpyxl.Workbook':
        """
        Create a new Excel workbook.
        
        Returns:
            OpenPyXL Workbook object
        """
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        return self.workbook
    
    def add_findings_sheet(
        self,
        findings: List[Dict[str, Any]],
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Add findings sheet to workbook.
        
        Args:
            findings: List of finding dictionaries
            sheet_name: Name for the sheet (defaults to config value)
        """
        sheet_name = sheet_name or self.config.sheet_findings
        self.worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Define headers
        headers = [
            'Evidence ID',
            'Event Name',
            'Event Time',
            'Resource Type',
            'Resource ID',
            'Priority',
            'Control Status',
            'Risk Score',
            'Risk Level',
            'Finding Title',
            'Finding Description',
            'Compliance Frameworks',
            'Remediation Available',
            'Remediation Action',
            'User Identity',
            'Source IP',
            'AWS Region',
            'AI Analyzed',
            'Model Used',
            'Collected At'
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            self._format_header(cell)
        
        # Write data
        for row, finding in enumerate(findings, start=2):
            self._write_finding_row(row, finding)
        
        # Auto-fit columns
        self._auto_fit_columns(self.worksheet, len(headers))
        
        # Freeze header row
        self.worksheet.freeze_panes = 'A2'
    
    def add_remediation_sheet(
        self,
        remediations: List[Dict[str, Any]],
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Add remediation sheet to workbook.
        
        Args:
            remediations: List of remediation action dictionaries
            sheet_name: Name for the sheet (defaults to config value)
        """
        sheet_name = sheet_name or self.config.sheet_remediation
        self.worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Define headers
        headers = [
            'Remediation ID',
            'Resource ID',
            'Resource Type',
            'Remediation Type',
            'Execution Mode',
            'Status',
            'Action Taken',
            'Result',
            'Error (if any)',
            'Triggered By',
            'Triggered At',
            'Completed At',
            'Success'
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            self._format_header(cell)
        
        # Write data
        for row, remediation in enumerate(remediations, start=2):
            self._write_remediation_row(row, remediation)
        
        # Auto-fit columns
        self._auto_fit_columns(self.worksheet, len(headers))
        
        # Freeze header row
        self.worksheet.freeze_panes = 'A2'
    
    def add_compliance_sheet(
        self,
        compliance_data: Dict[str, Any],
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Add compliance status sheet to workbook.
        
        Args:
            compliance_data: Compliance data by framework
            sheet_name: Name for the sheet (defaults to config value)
        """
        sheet_name = sheet_name or self.config.sheet_compliance
        self.worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Write title
        title_cell = self.worksheet.cell(row=1, column=1, value="Compliance Framework Status")
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        self.worksheet.merge_cells('A1:E1')
        
        # Write report period
        report_period = compliance_data.get('report_period', 'N/A')
        self.worksheet.cell(row=2, column=1, value=f"Report Period: {report_period}")
        self.worksheet.cell(row=2, column=1).font = Font(italic=True)
        
        # Define headers
        headers = [
            'Framework',
            'Version',
            'Total Controls',
            'Passed',
            'Failed',
            'Not Applicable',
            'Compliance %',
            'Status'
        ]
        
        start_row = 4
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=start_row, column=col, value=header)
            self._format_header(cell)
        
        # Write data for each framework
        frameworks = compliance_data.get('frameworks', [])
        for row, framework in enumerate(frameworks, start=start_row + 1):
            self._write_compliance_row(row, framework)
        
        # Auto-fit columns
        self._auto_fit_columns(self.worksheet, len(headers))
        
        # Freeze header row
        self.worksheet.freeze_panes = f'A{start_row + 1}'
    
    def add_summary_sheet(
        self,
        summary_data: Dict[str, Any],
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Add executive summary sheet to workbook.

        Args:
            summary_data: Summary data for the report
            sheet_name: Name for the sheet (defaults to config value)
        """
        sheet_name = sheet_name or self.config.sheet_summary
        self.worksheet = self.workbook.create_sheet(title=sheet_name)

        # Write title
        title_cell = self.worksheet.cell(row=1, column=1, value="Executive Summary")
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        self.worksheet.merge_cells('A1:E1')

        # Write report period
        report_period = summary_data.get('report_period', 'N/A')
        self.worksheet.cell(row=2, column=1, value=f"Report Period: {report_period}")
        self.worksheet.cell(row=2, column=1).font = Font(italic=True)

        # Overall Risk Score
        start_row = 4
        self.worksheet.cell(row=start_row, column=1, value="Overall Risk Score:")
        self.worksheet.cell(row=start_row, column=2, value=summary_data.get('overall_risk_score', 0))
        self.worksheet.cell(row=start_row, column=2).font = Font(bold=True, color='FF0000')

        # Total Evidence
        self.worksheet.cell(row=start_row + 1, column=1, value="Total Evidence Records:")
        self.worksheet.cell(row=start_row + 1, column=2, value=summary_data.get('total_evidence', 0))

        # Critical Findings
        critical_count = summary_data.get('critical_findings', 0)
        self.worksheet.cell(row=start_row + 2, column=1, value="Critical Findings:")
        self.worksheet.cell(row=start_row + 2, column=2, value=critical_count)
        if critical_count > 0:
            self.worksheet.cell(row=start_row + 2, column=2).font = Font(bold=True, color='FF0000')

        # High Findings
        high_count = summary_data.get('high_findings', 0)
        self.worksheet.cell(row=start_row + 3, column=1, value="High Findings:")
        self.worksheet.cell(row=start_row + 3, column=2, value=high_count)
        if high_count > 0:
            self.worksheet.cell(row=start_row + 3, column=2).font = Font(bold=True, color='FF9900')

        # Auto-remediation stats
        self.worksheet.cell(row=start_row + 5, column=1, value="Auto-Remediation:")
        self.worksheet.cell(row=start_row + 5, column=1).font = Font(bold=True)

        self.worksheet.cell(row=start_row + 6, column=1, value="  Successful:")
        self.worksheet.cell(row=start_row + 6, column=2, value=summary_data.get('successful_remediations', 0))

        self.worksheet.cell(row=start_row + 7, column=1, value="  Failed:")
        self.worksheet.cell(row=start_row + 7, column=2, value=summary_data.get('failed_remediations', 0))

        # Compliance Score
        self.worksheet.cell(row=start_row + 9, column=1, value="Compliance Score:")
        self.worksheet.cell(row=start_row + 9, column=2, value=f"{summary_data.get('compliance_score', 0)}%")

        # Auto-fit columns
        self._auto_fit_columns(self.worksheet, 2)

    def _format_header(self, cell: 'openpyxl.cell.Cell') -> None:
        """Apply header formatting to a cell."""
        cell.font = Font(**self.config.header_font)
        cell.fill = PatternFill(**self.config.header_fill)
        cell.alignment = Alignment(**self.config.header_alignment)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _write_finding_row(self, row: int, finding: Dict[str, Any]) -> None:
        """Write a single finding row to the worksheet."""
        data = [
            finding.get('evidence_id', 'N/A'),
            finding.get('event_name', 'N/A'),
            finding.get('event_time', 'N/A'),
            finding.get('resource_type', 'N/A'),
            finding.get('resource_id', 'N/A'),
            finding.get('priority', 'LOW'),
            finding.get('control_status', 'UNKNOWN'),
            finding.get('risk_score', 0),
            finding.get('risk_level', 'UNKNOWN'),
            finding.get('finding_title', 'N/A'),
            finding.get('finding_description', 'N/A'),
            self._format_list(finding.get('compliance_frameworks', [])),
            finding.get('remediation_available', 'No'),
            finding.get('remediation_action', 'N/A'),
            finding.get('user_identity', 'N/A'),
            finding.get('source_ip', 'N/A'),
            finding.get('aws_region', 'N/A'),
            finding.get('ai_analyzed', 'No'),
            finding.get('model_used', 'N/A'),
            finding.get('collected_at', 'N/A')
        ]

        for col, value in enumerate(data, start=1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            self._format_data_cell(cell, value, col)

    def _write_remediation_row(self, row: int, remediation: Dict[str, Any]) -> None:
        """Write a single remediation row to the worksheet."""
        data = [
            remediation.get('id', 'N/A'),
            remediation.get('resource_id', 'N/A'),
            remediation.get('resource_type', 'N/A'),
            remediation.get('remediation_type', 'N/A'),
            remediation.get('execution_mode', 'UNKNOWN'),
            remediation.get('status', 'PENDING'),
            remediation.get('action_taken', 'N/A'),
            remediation.get('result', 'N/A'),
            remediation.get('error', 'N/A'),
            remediation.get('triggered_by', 'N/A'),
            remediation.get('triggered_at', 'N/A'),
            remediation.get('completed_at', 'N/A'),
            remediation.get('success', False)
        ]

        for col, value in enumerate(data, start=1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            self._format_data_cell(cell, value, col)

    def _write_compliance_row(self, row: int, framework: Dict[str, Any]) -> None:
        """Write a single compliance framework row to the worksheet."""
        data = [
            framework.get('framework_name', 'N/A'),
            framework.get('version', 'N/A'),
            framework.get('total_controls', 0),
            framework.get('passed', 0),
            framework.get('failed', 0),
            framework.get('not_applicable', 0),
            f"{framework.get('compliance_percentage', 0)}%",
            framework.get('status', 'UNKNOWN')
        ]

        for col, value in enumerate(data, start=1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            self._format_data_cell(cell, value, col)

            # Color code compliance status
            if col == 8:  # Status column
                if value == 'COMPLIANT':
                    cell.font = Font(color='008000')
                elif value == 'NON_COMPLIANT':
                    cell.font = Font(color='FF0000', bold=True)

    def _format_data_cell(self, cell: 'openpyxl.cell.Cell', value: Any, col: int) -> None:
        """Apply formatting to a data cell."""
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Color code priority/status columns
        if col == 7:  # Control Status column
            if value == 'PASS':
                cell.font = Font(color='008000')
            elif value == 'FAIL':
                cell.font = Font(color='FF0000', bold=True)
        elif col == 6:  # Priority column
            if value == 'CRITICAL':
                cell.font = Font(color='FF0000', bold=True)
            elif value == 'HIGH':
                cell.font = Font(color='FF9900', bold=True)
            elif value == 'MEDIUM':
                cell.font = Font(color='FFCC00')

    def _format_list(self, items: list) -> str:
        """Format a list as a comma-separated string."""
        if isinstance(items, list):
            return ', '.join(str(item) for item in items)
        return str(items) if items else 'N/A'

    def _auto_fit_columns(self, worksheet: 'openpyxl.worksheet.Worksheet', num_columns: int) -> None:
        """Auto-fit column widths based on content."""
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width with some padding (max 50)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_workbook(self, file_path: str) -> None:
        """
        Save the workbook to a file.

        Args:
            file_path: Path to save the workbook
        """
        self.workbook.save(file_path)

    def get_workbook_bytes(self) -> bytes:
        """
        Get the workbook as bytes.

        Returns:
            Workbook as bytes
        """
        from io import BytesIO
        output = BytesIO()
        self.workbook.save(output)
        return output.getvalue()

    def generate_comprehensive_report(
        self,
        findings: List[Dict[str, Any]],
        remediations: List[Dict[str, Any]],
        compliance_data: Dict[str, Any],
        summary_data: Dict[str, Any]
    ) -> 'openpyxl.Workbook':
        """
        Generate a comprehensive Excel report with all sheets.

        Args:
            findings: List of finding dictionaries
            remediations: List of remediation action dictionaries
            compliance_data: Compliance data by framework
            summary_data: Executive summary data

        Returns:
            OpenPyXL Workbook object
        """
        # Create workbook
        self.create_workbook()

        # Add all sheets
        self.add_summary_sheet(summary_data)
        self.add_findings_sheet(findings)
        self.add_remediation_sheet(remediations)
        self.add_compliance_sheet(compliance_data)

        return self.workbook
